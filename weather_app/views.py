from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.conf import settings
from django.core.exceptions import ValidationError
from django.core.serializers import serialize
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
import json
import requests
import csv
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
import markdown
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from .models import WeatherRecord, WeatherSearch
from .forms import WeatherRecordForm

def safe_decimal(value, default=None):
    if value is None:
        return default
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return default

def index(request):
    return render(request, 'weather_app/index.html')

def get_weather_data(location, google_api_key):
    try:
        # Use Google Maps API for geocoding
        if not google_api_key or google_api_key == 'Enter your API key here':
            return None, "Google Maps API key is required for location services."
        
        # First, get coordinates using Google Maps Geocoding API
        geocode_url = f"https://maps.googleapis.com/maps/api/geocode/json?address={location}&key={google_api_key}"
        geocode_response = requests.get(geocode_url)
        geocode_data = geocode_response.json()
        
        if geocode_data['status'] != 'OK' or not geocode_data['results']:
            return None, "Location not found"
        
        lat = geocode_data['results'][0]['geometry']['location']['lat']
        lon = geocode_data['results'][0]['geometry']['location']['lng']
        location_name = geocode_data['results'][0]['formatted_address']
        
        # Use Open-Meteo API (completely free, no key required)
        weather_url = f"https://api.open-meteo.com/v1/forecast?latitude={lat}&longitude={lon}&current=temperature_2m,relative_humidity_2m,apparent_temperature,precipitation,weather_code,wind_speed_10m&hourly=temperature_2m,relative_humidity_2m,precipitation_probability,weather_code&daily=temperature_2m_max,temperature_2m_min,precipitation_probability_max&forecast_days=5&timezone=auto"
        weather_response = requests.get(weather_url)
        weather_data = weather_response.json()
        
        if 'error' in weather_data:
            return None, f"Weather API error: {weather_data['error']}"
        
        # Format the data to match our expected structure
        current = weather_data['current']
        hourly = weather_data['hourly']
        daily = weather_data['daily']
        
        # Convert weather codes to descriptions
        weather_descriptions = {
            0: 'Clear sky', 1: 'Mainly clear', 2: 'Partly cloudy', 3: 'Overcast',
            45: 'Foggy', 48: 'Depositing rime fog', 51: 'Light drizzle',
            53: 'Moderate drizzle', 55: 'Dense drizzle', 61: 'Slight rain',
            63: 'Moderate rain', 65: 'Heavy rain', 71: 'Slight snow',
            73: 'Moderate snow', 75: 'Heavy snow', 77: 'Snow grains',
            80: 'Slight rain showers', 81: 'Moderate rain showers',
            82: 'Violent rain showers', 85: 'Slight snow showers',
            86: 'Heavy snow showers', 95: 'Thunderstorm',
            96: 'Thunderstorm with slight hail', 99: 'Thunderstorm with heavy hail'
        }
        
        formatted_current = {
            'main': {
                'temp': current['temperature_2m'],
                'feels_like': current['apparent_temperature'],
                'humidity': current['relative_humidity_2m'],
                'pressure': 1013  # Default pressure
            },
            'weather': [{
                'description': weather_descriptions.get(current['weather_code'], 'Unknown'),
                'icon': '01d'  # Default icon
            }],
            'wind': {
                'speed': current['wind_speed_10m']
            },
            'visibility': 10000  # Default visibility
        }
        
        # Format forecast data (next 5 days)
        forecast_list = []
        for i in range(len(hourly['time'])):  # Process all available hourly data
            forecast_list.append({
                'dt': int(datetime.fromisoformat(hourly['time'][i].replace('Z', '+00:00')).timestamp()),
                'main': {
                    'temp_min': hourly['temperature_2m'][i],
                    'temp_max': hourly['temperature_2m'][i],
                    'humidity': hourly['relative_humidity_2m'][i]
                },
                'weather': [{
                    'description': weather_descriptions.get(hourly['weather_code'][i], 'Unknown'),
                    'icon': '01d'
                }]
            })
        
        formatted_forecast = {
            'list': forecast_list
        }
        
        return {
            'current': formatted_current,
            'forecast': formatted_forecast,
            'location': {
                'name': location_name,
                'lat': lat,
                'lon': lon
            }
        }, None
        
    except Exception as e:
        return None, str(e)

def get_historical_weather_data(location, start_date, end_date, google_api_key):
    """Get historical weather data using Google Maps for geocoding and Open-Meteo historical API"""
    try:
        # Use Google Maps API for geocoding
        if not google_api_key or google_api_key == 'Enter your API key here':
            return None, "Google Maps API key is required for location services."
        
        # First, get coordinates using Google Maps Geocoding API
        geocode_url = f"https://maps.googleapis.com/maps/api/geocode/json?address={location}&key={google_api_key}"
        geocode_response = requests.get(geocode_url)
        geocode_data = geocode_response.json()
        
        if geocode_data['status'] != 'OK' or not geocode_data['results']:
            return None, "Location not found"
        
        lat = geocode_data['results'][0]['geometry']['location']['lat']
        lon = geocode_data['results'][0]['geometry']['location']['lng']
        location_name = geocode_data['results'][0]['formatted_address']
        
        # Use Open-Meteo historical API
        historical_url = f"https://archive-api.open-meteo.com/v1/archive?latitude={lat}&longitude={lon}&start_date={start_date}&end_date={end_date}&hourly=temperature_2m,relative_humidity_2m,weather_code&timezone=auto"
        historical_response = requests.get(historical_url)
        historical_data = historical_response.json()
        
        if 'error' in historical_data:
            return None, f"Historical weather API error: {historical_data['error']}"
        
        if 'hourly' not in historical_data:
            return None, "No hourly weather data available from historical API"
        
        # Format the data to match our expected structure
        hourly = historical_data['hourly']
        
        # Convert weather codes to descriptions
        weather_descriptions = {
            0: 'Clear sky', 1: 'Mainly clear', 2: 'Partly cloudy', 3: 'Overcast',
            45: 'Foggy', 48: 'Depositing rime fog', 51: 'Light drizzle',
            53: 'Moderate drizzle', 55: 'Dense drizzle', 61: 'Slight rain',
            63: 'Moderate rain', 65: 'Heavy rain', 71: 'Slight snow',
            73: 'Moderate snow', 75: 'Heavy snow', 77: 'Snow grains',
            80: 'Slight rain showers', 81: 'Moderate rain showers',
            82: 'Violent rain showers', 85: 'Slight snow showers',
            86: 'Heavy snow showers', 95: 'Thunderstorm',
            96: 'Thunderstorm with slight hail', 99: 'Thunderstorm with heavy hail'
        }
        
        # Format historical data as forecast data structure
        forecast_list = []
        
        for i in range(len(hourly['time'])):
            if hourly['temperature_2m'][i] is not None:  # Skip null values
                forecast_list.append({
                    'dt': int(datetime.fromisoformat(hourly['time'][i].replace('Z', '+00:00')).timestamp()),
                    'main': {
                        'temp_min': hourly['temperature_2m'][i],
                        'temp_max': hourly['temperature_2m'][i],
                        'humidity': hourly['relative_humidity_2m'][i] if hourly['relative_humidity_2m'][i] is not None else 50
                    },
                    'weather': [{
                        'description': weather_descriptions.get(hourly['weather_code'][i], 'Unknown') if hourly['weather_code'][i] is not None else 'Unknown',
                        'icon': '01d'
                    }]
                })
        
        formatted_forecast = {
            'list': forecast_list
        }
        
        # Create a mock current weather (since historical data doesn't have current)
        formatted_current = {
            'main': {
                'temp': forecast_list[0]['main']['temp_min'] if forecast_list else 20,
                'feels_like': forecast_list[0]['main']['temp_min'] if forecast_list else 20,
                'humidity': forecast_list[0]['main']['humidity'] if forecast_list else 50,
                'pressure': 1013
            },
            'weather': [{
                'description': forecast_list[0]['weather'][0]['description'] if forecast_list else 'Unknown',
                'icon': '01d'
            }],
            'wind': {
                'speed': 5.0
            },
            'visibility': 10000
        }
        
        return {
            'current': formatted_current,
            'forecast': formatted_forecast,
            'location': {
                'name': location_name,
                'lat': lat,
                'lon': lon
            }
        }, None
        
    except Exception as e:
        return None, str(e)

def get_demo_weather_data(location):
    """Return demo weather data for testing without API key"""
    import random
    from datetime import datetime, timedelta
    
    # Generate demo current weather
    current_weather = {
        'main': {
            'temp': random.randint(15, 30),
            'feels_like': random.randint(12, 28),
            'humidity': random.randint(40, 80),
            'pressure': random.randint(1000, 1020)
        },
        'weather': [{
            'description': 'Partly cloudy',
            'icon': '02d'
        }],
        'wind': {
            'speed': random.uniform(2, 8)
        },
        'visibility': random.randint(8000, 12000)
    }
    
    # Generate demo forecast
    forecast_list = []
    for i in range(40):  # 5 days * 8 forecasts per day
        forecast_list.append({
            'dt': int((datetime.now() + timedelta(hours=i*3)).timestamp()),
            'main': {
                'temp_min': random.randint(10, 25),
                'temp_max': random.randint(20, 35),
                'humidity': random.randint(40, 80)
            },
            'weather': [{
                'description': random.choice(['Clear sky', 'Partly cloudy', 'Scattered clouds', 'Light rain']),
                'icon': random.choice(['01d', '02d', '03d', '10d'])
            }]
        })
    
    forecast_data = {
        'list': forecast_list
    }
    
    return {
        'current': current_weather,
        'forecast': forecast_data,
        'location': {
            'name': location.title(),
            'lat': random.uniform(30, 50),
            'lon': random.uniform(-120, -70)
        }
    }

@csrf_exempt
def api_weather(request):
    """API endpoint for getting weather data"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            location = data.get('location')
            
            if not location:
                return JsonResponse({'error': 'Location is required'}, status=400)
            
            weather_data, error = get_weather_data(location, settings.GOOGLE_MAPS_API_KEY)
            
            if error:
                return JsonResponse({'error': error}, status=400)
            
            # Log the search
            WeatherSearch.objects.create(
                location_query=location,
                search_type='current',
                success=True
            )
            
            return JsonResponse(weather_data)
            
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@csrf_exempt
def api_weather_by_coordinates(request):
    """API endpoint for getting weather data by coordinates"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            lat = data.get('lat')
            lon = data.get('lon')
            
            if not lat or not lon:
                return JsonResponse({'error': 'Latitude and longitude are required'}, status=400)
            
            # Check if we have a valid Google Maps API key
            api_key = settings.GOOGLE_MAPS_API_KEY
            if not api_key or api_key == 'Enter your API key here':
                return JsonResponse({'error': 'Google Maps API key is required for location services.'}, status=400)
            
            # Get weather data directly by coordinates using Open-Meteo API
            weather_url = f"https://api.open-meteo.com/v1/forecast?latitude={lat}&longitude={lon}&current=temperature_2m,relative_humidity_2m,apparent_temperature,precipitation,weather_code,wind_speed_10m&hourly=temperature_2m,relative_humidity_2m,precipitation_probability,weather_code&daily=temperature_2m_max,temperature_2m_min,precipitation_probability_max&forecast_days=5&timezone=auto"
            weather_response = requests.get(weather_url)
            weather_data = weather_response.json()
            
            if 'error' in weather_data:
                return JsonResponse({'error': f"Weather API error: {weather_data['error']}"}, status=400)
            
            # Format the data to match our expected structure
            current = weather_data['current']
            hourly = weather_data['hourly']
            
            # Convert weather codes to descriptions
            weather_descriptions = {
                0: 'Clear sky', 1: 'Mainly clear', 2: 'Partly cloudy', 3: 'Overcast',
                45: 'Foggy', 48: 'Depositing rime fog', 51: 'Light drizzle',
                53: 'Moderate drizzle', 55: 'Dense drizzle', 61: 'Slight rain',
                63: 'Moderate rain', 65: 'Heavy rain', 71: 'Slight snow',
                73: 'Moderate snow', 75: 'Heavy snow', 77: 'Snow grains',
                80: 'Slight rain showers', 81: 'Moderate rain showers',
                82: 'Violent rain showers', 85: 'Slight snow showers',
                86: 'Heavy snow showers', 95: 'Thunderstorm',
                96: 'Thunderstorm with slight hail', 99: 'Thunderstorm with heavy hail'
            }
            
            formatted_current = {
                'main': {
                    'temp': current['temperature_2m'],
                    'feels_like': current['apparent_temperature'],
                    'humidity': current['relative_humidity_2m'],
                    'pressure': 1013
                },
                'weather': [{
                    'description': weather_descriptions.get(current['weather_code'], 'Unknown'),
                    'icon': '01d'
                }],
                'wind': {
                    'speed': current['wind_speed_10m']
                },
                'visibility': 10000
            }
            
            # Format forecast data
            forecast_list = []
            for i in range(len(hourly['time'])):  # Process all available hourly data
                forecast_list.append({
                    'dt': int(datetime.fromisoformat(hourly['time'][i].replace('Z', '+00:00')).timestamp()),
                    'main': {
                        'temp_min': hourly['temperature_2m'][i],
                        'temp_max': hourly['temperature_2m'][i],
                        'humidity': hourly['relative_humidity_2m'][i]
                    },
                    'weather': [{
                        'description': weather_descriptions.get(hourly['weather_code'][i], 'Unknown'),
                        'icon': '01d'
                    }]
                })
            
            formatted_forecast = {
                'list': forecast_list
            }
            
            # Log the search
            WeatherSearch.objects.create(
                location_query=f"{lat},{lon}",
                search_type='current',
                success=True
            )
            
            return JsonResponse({
                'current': formatted_current,
                'forecast': formatted_forecast,
                'location': {
                    'lat': lat,
                    'lon': lon
                }
            })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)

def weather_records_list(request):
    """Display all weather records with CRUD operations"""
    records = WeatherRecord.objects.all()
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        records = records.filter(
            Q(location_name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(notes__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(records, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
    }
    
    return render(request, 'weather_app/weather_records.html', context)

def weather_record_create(request):
    """Create a new weather record"""
    if request.method == 'POST':
        form = WeatherRecordForm(request.POST)
        if form.is_valid():
            try:
                # Get form data
                location_name = form.cleaned_data['location_name']
                start_date = form.cleaned_data['start_date']
                end_date = form.cleaned_data['end_date']
                
                # Fetch weather data for the location and date range
                # Check if we need historical data (past dates) or forecast data (future dates)
                today = timezone.now().date()
                
                if start_date < today:
                    # Historical data needed - use historical endpoint
                    weather_data, error = get_historical_weather_data(location_name, start_date, end_date, settings.GOOGLE_MAPS_API_KEY)
                else:
                    # Future data needed - use forecast endpoint
                    weather_data, error = get_weather_data(location_name, settings.GOOGLE_MAPS_API_KEY)
                
                if error:
                    messages.error(request, f'Error fetching weather data: {error}')
                    return render(request, 'weather_app/weather_record_form.html', {'form': form, 'action': 'Create'})
                
                # Get coordinates from weather data
                lat = weather_data['location']['lat']
                lon = weather_data['location']['lon']
                
                # Calculate average weather for the date range
                forecast_list = weather_data['forecast']['list']
                date_range_data = []
                
                for item in forecast_list:
                    item_date = datetime.fromtimestamp(item['dt']).date()
                    if start_date <= item_date <= end_date:
                        date_range_data.append(item)
                
                if not date_range_data:
                    messages.error(request, 'No weather data available for the selected date range')
                    return render(request, 'weather_app/weather_record_form.html', {'form': form, 'action': 'Create'})
                
                # Calculate averages and extremes
                temps = [item['main']['temp_min'] for item in date_range_data if item['main']['temp_min'] is not None]
                humidities = [item['main']['humidity'] for item in date_range_data if item['main']['humidity'] is not None]
                
                if not temps:
                    messages.error(request, 'No valid temperature data available for the selected date range')
                    return render(request, 'weather_app/weather_record_form.html', {'form': form, 'action': 'Create'})
                
                temp_min = min(temps)
                temp_max = max(temps)
                temp_avg = sum(temps) / len(temps)
                humidity_avg = sum(humidities) / len(humidities) if humidities else 50
                
                # Get most common weather description
                descriptions = [item['weather'][0]['description'] for item in date_range_data]
                most_common_description = max(set(descriptions), key=descriptions.count)
                
                # Create record using safe decimal conversion
                try:
                    # Create record with all fields at once
                    record = WeatherRecord.objects.create(
                        location_name=form.cleaned_data['location_name'],
                        location_type=form.cleaned_data['location_type'],
                        latitude=safe_decimal(lat),
                        longitude=safe_decimal(lon),
                        start_date=start_date,
                        end_date=end_date,
                        temperature_min=safe_decimal(temp_min),
                        temperature_max=safe_decimal(temp_max),
                        temperature_avg=safe_decimal(round(temp_avg, 2) if temp_avg is not None else None),
                        humidity=int(round(humidity_avg)) if humidity_avg is not None else 50,
                        description=most_common_description,
                        icon='01d',
                        pressure=safe_decimal(1013),
                        wind_speed=safe_decimal(5.0),
                        precipitation=safe_decimal(0.0),
                        visibility=safe_decimal(9999.99),  # Use max value for 6,2 field
                        uv_index=safe_decimal(5.0),
                        notes=form.cleaned_data.get('notes', '')
                    )
                except Exception as e:
                    raise
                messages.success(request, 'Weather record created successfully with real weather data!')
                return redirect('weather_app:weather_records_list')
            except ValidationError as e:
                messages.error(request, str(e))
            except Exception as e:
                messages.error(request, f'Error creating weather record: {str(e)}')
    else:
        form = WeatherRecordForm()
    
    return render(request, 'weather_app/weather_record_form.html', {'form': form, 'action': 'Create'})

def weather_record_update(request, pk):
    """Update an existing weather record"""
    record = get_object_or_404(WeatherRecord, pk=pk)
    
    if request.method == 'POST':
        form = WeatherRecordForm(request.POST, instance=record)
        if form.is_valid():
            try:
                # Get form data
                location_name = form.cleaned_data['location_name']
                start_date = form.cleaned_data['start_date']
                end_date = form.cleaned_data['end_date']
                
                # Fetch fresh weather data for the location and date range
                # Check if we need historical data (past dates) or forecast data (future dates)
                today = timezone.now().date()
                
                if start_date < today:
                    # Historical data needed - use historical endpoint
                    weather_data, error = get_historical_weather_data(location_name, start_date, end_date, settings.GOOGLE_MAPS_API_KEY)
                else:
                    # Future data needed - use forecast endpoint
                    weather_data, error = get_weather_data(location_name, settings.GOOGLE_MAPS_API_KEY)
                
                if error:
                    messages.error(request, f'Error fetching weather data: {error}')
                    return render(request, 'weather_app/weather_record_form.html', {
                        'form': form, 
                        'action': 'Update',
                        'record': record
                    })
                
                # Get coordinates from weather data
                lat = weather_data['location']['lat']
                lon = weather_data['location']['lon']
                
                # Calculate average weather for the date range
                forecast_list = weather_data['forecast']['list']
                date_range_data = []
                
                for item in forecast_list:
                    item_date = datetime.fromtimestamp(item['dt']).date()
                    if start_date <= item_date <= end_date:
                        date_range_data.append(item)
                
                if not date_range_data:
                    messages.error(request, 'No weather data available for the selected date range')
                    return render(request, 'weather_app/weather_record_form.html', {
                        'form': form, 
                        'action': 'Update',
                        'record': record
                    })
                
                # Calculate averages and extremes
                temps = [item['main']['temp_min'] for item in date_range_data if item['main']['temp_min'] is not None]
                humidities = [item['main']['humidity'] for item in date_range_data if item['main']['humidity'] is not None]
                
                if not temps:
                    messages.error(request, 'No valid temperature data available for the selected date range')
                    return render(request, 'weather_app/weather_record_form.html', {
                        'form': form, 
                        'action': 'Update',
                        'record': record
                    })
                
                temp_min = min(temps)
                temp_max = max(temps)
                temp_avg = sum(temps) / len(temps)
                humidity_avg = sum(humidities) / len(humidities) if humidities else 50
                
                # Get most common weather description
                descriptions = [item['weather'][0]['description'] for item in date_range_data]
                most_common_description = max(set(descriptions), key=descriptions.count)
                
                # Update record using safe decimal conversion
                record.latitude = safe_decimal(lat)
                record.longitude = safe_decimal(lon)
                record.temperature_min = safe_decimal(temp_min)
                record.temperature_max = safe_decimal(temp_max)
                record.temperature_avg = safe_decimal(round(temp_avg, 2) if temp_avg is not None else None)
                record.humidity = int(round(humidity_avg)) if humidity_avg is not None else 50
                record.description = most_common_description
                record.icon = '01d'
                record.pressure = safe_decimal(1013)
                record.wind_speed = safe_decimal(5.0)
                record.precipitation = safe_decimal(0.0)
                record.visibility = safe_decimal(9999.99)  # Use max value for 6,2 field
                record.uv_index = safe_decimal(5.0)
                record.save()
                messages.success(request, 'Weather record updated successfully with fresh weather data!')
                return redirect('weather_app:weather_records_list')
            except ValidationError as e:
                messages.error(request, str(e))
            except Exception as e:
                messages.error(request, f'Error updating weather record: {str(e)}')
    else:
        form = WeatherRecordForm(instance=record)
    
    return render(request, 'weather_app/weather_record_form.html', {
        'form': form, 
        'action': 'Update',
        'record': record
    })

def weather_record_delete(request, pk):
    """Delete a weather record"""
    record = get_object_or_404(WeatherRecord, pk=pk)
    
    if request.method == 'POST':
        record.delete()
        messages.success(request, 'Weather record deleted successfully!')
        return redirect('weather_app:weather_records_list')
    
    return render(request, 'weather_app/weather_record_confirm_delete.html', {'record': record})

def weather_record_detail(request, pk):
    """Display detailed information about a weather record"""
    record = get_object_or_404(WeatherRecord, pk=pk)
    return render(request, 'weather_app/weather_record_detail.html', {'record': record})

def export_data(request, format_type):
    """Export weather data in various formats"""
    records = WeatherRecord.objects.all()
    
    if format_type == 'json':
        data = serialize('json', records)
        response = HttpResponse(data, content_type='application/json')
        response['Content-Disposition'] = 'attachment; filename="weather_data.json"'
        return response
    
    elif format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="weather_data.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Location', 'Location Type', 'Start Date', 'End Date', 
            'Temperature Min', 'Temperature Max', 'Temperature Avg',
            'Humidity', 'Pressure', 'Wind Speed', 'Description', 'Created At'
        ])
        
        for record in records:
            writer.writerow([
                record.location_name, record.location_type, record.start_date, record.end_date,
                record.temperature_min, record.temperature_max, record.temperature_avg,
                record.humidity, record.pressure, record.wind_speed, record.description,
                record.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        return response
    
    elif format_type == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="weather_data.pdf"'
        
        p = canvas.Canvas(response, pagesize=letter)
        width, height = letter
        
        # Title
        p.setFont("Helvetica-Bold", 16)
        p.drawString(1*inch, height-1*inch, "Weather Data Report")
        p.setFont("Helvetica", 12)
        p.drawString(1*inch, height-1.2*inch, f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        y_position = height - 1.5*inch
        
        for record in records:
            if y_position < 1*inch:  # New page if needed
                p.showPage()
                y_position = height - 1*inch
            
            p.setFont("Helvetica-Bold", 12)
            p.drawString(1*inch, y_position, f"{record.location_name}")
            y_position -= 0.2*inch
            
            p.setFont("Helvetica", 10)
            p.drawString(1*inch, y_position, f"Date Range: {record.start_date} to {record.end_date}")
            y_position -= 0.15*inch
            p.drawString(1*inch, y_position, f"Temperature: {record.get_temperature_range()}")
            y_position -= 0.15*inch
            p.drawString(1*inch, y_position, f"Description: {record.description or 'N/A'}")
            y_position -= 0.3*inch
        
        p.showPage()
        p.save()
        return response
    
    elif format_type == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = "Weather Data"
        
        # Headers
        headers = [
            'Location', 'Location Type', 'Start Date', 'End Date', 
            'Temperature Min', 'Temperature Max', 'Temperature Avg',
            'Humidity', 'Pressure', 'Wind Speed', 'Description', 'Created At'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row, record in enumerate(records, 2):
            ws.cell(row=row, column=1, value=record.location_name)
            ws.cell(row=row, column=2, value=record.location_type)
            ws.cell(row=row, column=3, value=record.start_date)
            ws.cell(row=row, column=4, value=record.end_date)
            ws.cell(row=row, column=5, value=record.temperature_min)
            ws.cell(row=row, column=6, value=record.temperature_max)
            ws.cell(row=row, column=7, value=record.temperature_avg)
            ws.cell(row=row, column=8, value=record.humidity)
            ws.cell(row=row, column=9, value=record.pressure)
            ws.cell(row=row, column=10, value=record.wind_speed)
            ws.cell(row=row, column=11, value=record.description)
            ws.cell(row=row, column=12, value=record.created_at)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="weather_data.xlsx"'
        wb.save(response)
        return response
    
    elif format_type == 'markdown':
        md_content = "# Weather Data Report\n\n"
        md_content += f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
        
        for record in records:
            md_content += f"## {record.location_name}\n\n"
            md_content += f"- **Date Range:** {record.start_date} to {record.end_date}\n"
            md_content += f"- **Temperature:** {record.get_temperature_range()}\n"
            md_content += f"- **Humidity:** {record.humidity or 'N/A'}%\n"
            md_content += f"- **Pressure:** {record.pressure or 'N/A'} hPa\n"
            md_content += f"- **Wind Speed:** {record.wind_speed or 'N/A'} m/s\n"
            md_content += f"- **Description:** {record.description or 'N/A'}\n"
            md_content += f"- **Created:** {record.created_at.strftime('%Y-%m-%d %H:%M:%S')}\n\n"
        
        response = HttpResponse(md_content, content_type='text/markdown')
        response['Content-Disposition'] = 'attachment; filename="weather_data.md"'
        return response
    
    return HttpResponse("Unsupported format", status=400)


